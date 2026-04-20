"""Tests for parser.py."""
from pathlib import Path
from unittest.mock import MagicMock, patch

import numpy as np
import pandas as pd
import pytest

from augean import parser
from augean.parser import (
    _split_clinical_indication,
    extract_label_scan,
    extract_named_cells,
    extract_named_cells_multi,
    extract_tabular,
    merge_dataframes,
    parse_workbook,
)

WORKBOOKS_DIR = Path(__file__).parent / "test_data" / "workbooks"


# ---------------------------------------------------------------------------
# Unit tests for _split_clinical_indication
# ---------------------------------------------------------------------------

class TestSplitClinicalIndication:
    def test_single_indication(self):
        names, codes = _split_clinical_indication("R208.1_Inherited breast cancer")
        assert codes == "R208.1"
        assert names == "Inherited breast cancer"

    def test_multiple_indications(self):
        names, codes = _split_clinical_indication(
            "R208.1_Inherited breast cancer;R207.1_Ovarian cancer"
        )
        assert codes == "R208.1;R207.1"
        assert names == "Inherited breast cancer;Ovarian cancer"

    def test_none_input(self):
        names, codes = _split_clinical_indication(None)
        assert names == ""
        assert codes == ""

    def test_no_underscore(self):
        names, codes = _split_clinical_indication("Bare condition")
        assert codes == "Bare condition"
        assert names == "Bare condition"


# ---------------------------------------------------------------------------
# extract_named_cells
# ---------------------------------------------------------------------------

class TestExtractNamedCells:
    def _make_sheet_with_cells(self, cell_values: dict):
        """Build a mock sheet where cell_values = {addr: value}."""
        sheet = MagicMock()

        def getitem(key):
            # Column iteration (single letter key)
            if isinstance(key, str) and len(key) == 1 and key.isalpha():
                col_data = cell_values.get(f"__col_{key}__", [])
                cells = []
                for i, val in enumerate(col_data, start=1):
                    mc = MagicMock()
                    mc.value = val
                    mc.row = i
                    cells.append(mc)
                return cells
            # Regular cell
            mc = MagicMock()
            mc.value = cell_values.get(key)
            return mc

        sheet.__getitem__ = lambda s, k: getitem(k)
        return sheet

    def test_extracts_simple_fields(self):
        wb = MagicMock()
        sheet = self._make_sheet_with_cells({"B1": "SAMPLE123", "F2": "Panel1"})
        wb.__getitem__ = lambda s, k: sheet

        config = {
            "fields": [
                {"db_column": "sample_id", "cell": "B1"},
                {"db_column": "panel",     "cell": "F2"},
            ]
        }
        df = extract_named_cells(wb, "summary", config)
        assert df["sample_id"][0] == "SAMPLE123"
        assert df["panel"][0] == "Panel1"

    def test_sentinel_scan(self):
        wb = MagicMock()
        sheet = self._make_sheet_with_cells({
            "__col_A__": [None, "Reference:", None],
            "B2": "GRCh38",
        })
        wb.__getitem__ = lambda s, k: sheet

        config = {
            "fields": [
                {
                    "db_column": "ref_genome",
                    "extraction": "sentinel_scan",
                    "scan_column": "A",
                    "sentinel_value": "Reference:",
                    "value_column": "B",
                }
            ]
        }
        df = extract_named_cells(wb, "summary", config)
        assert df["ref_genome"][0] == "GRCh38"

    def test_sentinel_scan_returns_none_when_not_found(self):
        wb = MagicMock()
        sheet = self._make_sheet_with_cells({
            "__col_A__": [None, None],
        })
        wb.__getitem__ = lambda s, k: sheet

        config = {
            "fields": [
                {
                    "db_column": "ref_genome",
                    "extraction": "sentinel_scan",
                    "scan_column": "A",
                    "sentinel_value": "Reference:",
                    "value_column": "B",
                }
            ]
        }
        df = extract_named_cells(wb, "summary", config)
        assert df["ref_genome"][0] is None

    def test_clinical_indication_split(self):
        wb = MagicMock()
        sheet = self._make_sheet_with_cells({
            "F1": "R208.1_Inherited breast cancer;R207.1_Ovarian cancer"
        })
        wb.__getitem__ = lambda s, k: sheet

        config = {
            "fields": [
                {"db_column": "clinical_indication", "cell": "F1",
                 "parse": "clinical_indication_split"}
            ]
        }
        df = extract_named_cells(wb, "summary", config)
        assert "preferred_condition_name" in df.columns
        assert "r_code" in df.columns
        assert df["preferred_condition_name"][0] == "Inherited breast cancer;Ovarian cancer"
        assert df["r_code"][0] == "R208.1;R207.1"
        assert "clinical_indication" not in df.columns

    def test_empty_cell_returns_none(self):
        wb = MagicMock()
        sheet = self._make_sheet_with_cells({"G22": None})
        wb.__getitem__ = lambda s, k: sheet

        config = {
            "fields": [
                {"db_column": "date_last_evaluated", "cell": "G22", "type": "date"}
            ]
        }
        df = extract_named_cells(wb, "summary", config)
        assert df["date_last_evaluated"][0] is None

    def test_sample_id_split(self):
        wb = MagicMock()
        sheet = self._make_sheet_with_cells({
            "B1": "100033006-22363S0007-23NGSHO1-8128-M-96527893"
        })
        wb.__getitem__ = lambda s, k: sheet

        config = {
            "fields": [
                {"db_column": "sample_id", "cell": "B1", "parse": "sample_id_split"}
            ]
        }
        df = extract_named_cells(wb, "summary", config)
        assert df["instrument_id"][0] == "100033006"
        assert df["specimen_id"][0] == "22363S0007"
        assert df["batch_id"][0] == "23NGSHO1"
        assert df["test_code"][0] == "8128"
        assert df["probeset_id"][0] == "96527893"
        assert "sample_id" not in df.columns


# ---------------------------------------------------------------------------
# extract_label_scan
# ---------------------------------------------------------------------------

class TestExtractLabelScan:
    def _make_wb(self, col_a: list, col_b_map: dict):
        wb = MagicMock()
        sheet = MagicMock()

        def getitem(key):
            if key == "A":
                cells = []
                for i, val in enumerate(col_a, start=1):
                    mc = MagicMock()
                    mc.value = val
                    mc.row = i
                    cells.append(mc)
                return cells
            # B{row} access
            mc = MagicMock()
            mc.value = col_b_map.get(key)
            return mc

        sheet.__getitem__ = lambda s, k: getitem(k)
        wb.__getitem__ = lambda s, k: sheet
        return wb

    def test_extracts_fields_by_label(self):
        wb = self._make_wb(
            ["Sample ID", None, "M-code", "Subpanel analysed"],
            {"B1": "SAMPLE-001", "B3": "M87", "B4": "Myeloid"},
        )
        config = {
            "scan_column": "A",
            "value_column": "B",
            "fields": [
                {"db_column": "sample_id", "label": "Sample ID"},
                {"db_column": "panel",     "label": "M-code"},
                {"db_column": "preferred_condition_name", "label": "Subpanel analysed"},
            ],
        }
        df = extract_label_scan(wb, "summary", config)
        assert df["sample_id"][0] == "SAMPLE-001"
        assert df["panel"][0] == "M87"
        assert df["preferred_condition_name"][0] == "Myeloid"

    def test_missing_label_returns_none(self):
        wb = self._make_wb(["Sample ID"], {"B1": "S1"})
        config = {
            "scan_column": "A",
            "value_column": "B",
            "fields": [
                {"db_column": "ref_genome", "label": "Reference:"},
            ],
        }
        df = extract_label_scan(wb, "summary", config)
        assert df["ref_genome"][0] is None

    def test_sample_id_split(self):
        wb = self._make_wb(
            ["Sample ID", "M-code"],
            {"B1": "100033006-22363S0007-23NGSHO1-8128-M-96527893", "B2": "M87"},
        )
        config = {
            "scan_column": "A",
            "value_column": "B",
            "fields": [
                {"db_column": "sample_id", "label": "Sample ID", "parse": "sample_id_split"},
                {"db_column": "panel", "label": "M-code"},
            ],
        }
        df = extract_label_scan(wb, "summary", config)
        assert df["instrument_id"][0] == "100033006"
        assert df["specimen_id"][0] == "22363S0007"
        assert df["batch_id"][0] == "23NGSHO1"
        assert df["test_code"][0] == "8128"
        assert df["probeset_id"][0] == "96527893"
        assert df["panel"][0] == "M87"
        assert "sample_id" not in df.columns


# ---------------------------------------------------------------------------
# extract_named_cells_multi
# ---------------------------------------------------------------------------

class TestExtractNamedCellsMulti:
    def test_extracts_one_row_per_matching_sheet(self):
        wb = MagicMock()
        wb.sheetnames = ["summary", "interpret_1", "interpret_2", "included"]

        cells = {
            "interpret_1": {"C3": "NM_001.1:c.100A>T", "C26": "Pathogenic"},
            "interpret_2": {"C3": "NM_002.2:c.200G>C", "C26": "Likely pathogenic"},
        }

        def get_sheet(name):
            s = MagicMock()
            sheet_cells = cells.get(name, {})
            s.__getitem__ = lambda self, k: MagicMock(value=sheet_cells.get(k))
            return s

        wb.__getitem__ = lambda s, k: get_sheet(k)

        config = {
            "sheet_pattern": "^interpret",
            "fields": [
                {"db_column": "hgvsc",                   "cell": "C3"},
                {"db_column": "germline_classification", "cell": "C26"},
            ],
        }
        df = extract_named_cells_multi(wb, config)
        assert len(df) == 2
        assert list(df["hgvsc"]) == ["NM_001.1:c.100A>T", "NM_002.2:c.200G>C"]
        assert list(df["germline_classification"]) == ["Pathogenic", "Likely pathogenic"]

    def test_no_matching_sheets_returns_empty(self):
        wb = MagicMock()
        wb.sheetnames = ["summary", "included"]
        config = {
            "sheet_pattern": "^interpret",
            "fields": [{"db_column": "hgvsc", "cell": "C3"}],
        }
        df = extract_named_cells_multi(wb, config)
        assert df.empty


# ---------------------------------------------------------------------------
# extract_tabular (with real workbook)
# ---------------------------------------------------------------------------

class TestExtractTabular:
    def test_extracts_rd_dias_included(self, rd_dias_cuh_workbook, rd_dias_cuh_path, rd_dias_config):
        sheet_config = rd_dias_config["sheets"]["included"]
        df = extract_tabular(rd_dias_cuh_workbook, "included", sheet_config, rd_dias_cuh_path)

        assert "hgvsc" in df.columns
        assert "chromosome" in df.columns
        assert "local_id" in df.columns
        assert "linking_id" in df.columns
        assert len(df) >= 1

    def test_interpreted_lowercased(self, rd_dias_cuh_workbook, rd_dias_cuh_path, rd_dias_config):
        sheet_config = rd_dias_config["sheets"]["included"]
        df = extract_tabular(rd_dias_cuh_workbook, "included", sheet_config, rd_dias_cuh_path)

        assert df["interpreted"].isin(["yes", "no"]).all()

    def test_local_id_unique(self, rd_dias_cuh_workbook, rd_dias_cuh_path, rd_dias_config):
        sheet_config = rd_dias_config["sheets"]["included"]
        df = extract_tabular(rd_dias_cuh_workbook, "included", sheet_config, rd_dias_cuh_path)

        assert df["local_id"].nunique() == len(df)

    def test_linking_id_equals_local_id(self, rd_dias_cuh_workbook, rd_dias_cuh_path, rd_dias_config):
        sheet_config = rd_dias_config["sheets"]["included"]
        df = extract_tabular(rd_dias_cuh_workbook, "included", sheet_config, rd_dias_cuh_path)

        assert (df["local_id"] == df["linking_id"]).all()


# ---------------------------------------------------------------------------
# merge_dataframes
# ---------------------------------------------------------------------------

class TestMergeDataframes:
    def test_cross_join_then_left_join(self):
        summary = pd.DataFrame([{"sample_id": "S1", "panel": "P1"}])
        included = pd.DataFrame([
            {"hgvsc": "NM_1:c.100A>T", "chromosome": 1},
            {"hgvsc": "NM_2:c.200G>C", "chromosome": 2},
        ])
        interpret = pd.DataFrame([
            {"hgvsc": "NM_1:c.100A>T", "germline_classification": "Pathogenic"},
        ])
        merge_config = {
            "summary_x_included": {"how": "cross"},
            "included_join_interpret": {"on": "hgvsc", "how": "left"},
        }
        df = merge_dataframes(summary, included, interpret, merge_config)

        assert len(df) == 2
        assert "sample_id" in df.columns
        assert "germline_classification" in df.columns
        # Second variant has no interpret row → NaN classification
        row2 = df[df["hgvsc"] == "NM_2:c.200G>C"].iloc[0]
        assert pd.isna(row2["germline_classification"])

    def test_no_interpret_skips_join(self):
        summary = pd.DataFrame([{"sample_id": "S1"}])
        included = pd.DataFrame([{"hgvsc": "NM_1:c.100A>T"}])
        merge_config = {"summary_x_included": {"how": "cross"}}
        df = merge_dataframes(summary, included, None, merge_config)
        assert len(df) == 1
        assert "sample_id" in df.columns


# ---------------------------------------------------------------------------
# Integration: parse_workbook
# ---------------------------------------------------------------------------

class TestParseWorkbook:
    def test_rd_dias_parse_returns_dataframe(self, rd_dias_cuh_workbook, rd_dias_cuh_path, rd_dias_config):
        df = parse_workbook(rd_dias_cuh_workbook, rd_dias_config, rd_dias_cuh_path)

        assert isinstance(df, pd.DataFrame)
        assert len(df) >= 1
        assert "hgvsc" in df.columns
        assert "germline_classification" in df.columns
        assert "specimen_id" in df.columns
        assert "institution" in df.columns

    def test_rd_dias_constant_fields(self, rd_dias_cuh_workbook, rd_dias_cuh_path, rd_dias_config):
        df = parse_workbook(rd_dias_cuh_workbook, rd_dias_config, rd_dias_cuh_path)

        assert df["allele_origin"].unique()[0] == "germline"
        assert df["collection_method"].unique()[0] == "clinical testing"
        assert df["affected_status"].unique()[0] == "yes"

    def test_rd_dias_summary_fields(self, rd_dias_cuh_workbook, rd_dias_cuh_path, rd_dias_config):
        df = parse_workbook(rd_dias_cuh_workbook, rd_dias_config, rd_dias_cuh_path)

        assert df["ref_genome"][0] is not None
        assert "preferred_condition_name" in df.columns
        assert "r_code" in df.columns

    def test_haemonc_parse_returns_dataframe(self, haemonc_workbook, haemonc_path, haemonc_config):
        df = parse_workbook(haemonc_workbook, haemonc_config, haemonc_path)

        assert isinstance(df, pd.DataFrame)
        assert len(df) > 0
        assert "hgvsc" in df.columns
        assert "specimen_id" in df.columns
        assert "allele_origin" in df.columns

    def test_haemonc_allele_origin_somatic(self, haemonc_workbook, haemonc_path, haemonc_config):
        df = parse_workbook(haemonc_workbook, haemonc_config, haemonc_path)

        assert df["allele_origin"].unique()[0] == "somatic"

    def test_haemonc_variant_category_column_present(self, haemonc_workbook, haemonc_path, haemonc_config):
        df = parse_workbook(haemonc_workbook, haemonc_config, haemonc_path)

        assert "variant_category" in df.columns
        assert df["variant_category"].notna().all()

    def test_haemonc_allele_origin_still_somatic_after_pindel(self, haemonc_workbook, haemonc_path, haemonc_config):
        """Regression: top-level constant_fields broadcast to all rows including pindel."""
        df = parse_workbook(haemonc_workbook, haemonc_config, haemonc_path)

        assert (df["allele_origin"] == "somatic").all()


# ---------------------------------------------------------------------------
# pindel sheet extraction (TDD — written before implementation)
# ---------------------------------------------------------------------------

_MINIMAL_TWO_SHEET_CONFIG = {
    "sheets": {
        "included": {
            "extraction_type": "tabular",
            "constant_fields": {"variant_category": "included"},
            "columns": [{"source": "CHROM", "db_column": "chromosome"}],
            "generated_columns": [],
        },
        "pindel": {
            "extraction_type": "tabular",
            "constant_fields": {"variant_category": "pindel"},
            "columns": [{"source": "CHROM", "db_column": "chromosome"}],
            "generated_columns": [],
        },
    },
    "merge_strategy": {"summary_x_included": {"how": "cross"}},
    "constant_fields": {"allele_origin": "somatic"},
}


class TestExtractTabularConstantFields:
    """sheet-level constant_fields in extract_tabular."""

    def test_constant_fields_absent_no_extra_column(self, rd_dias_cuh_workbook, rd_dias_cuh_path, rd_dias_config):
        """Baseline: config without constant_fields produces no variant_category column."""
        sheet_config = rd_dias_config["sheets"]["included"]
        assert "constant_fields" not in sheet_config
        df = extract_tabular(rd_dias_cuh_workbook, "included", sheet_config, rd_dias_cuh_path)
        assert "variant_category" not in df.columns

    def test_constant_fields_stamped_on_all_rows(self, rd_dias_cuh_workbook, rd_dias_cuh_path, rd_dias_config):
        """Sheet-level constant_fields adds the column with the given value on every row."""
        sheet_config = {**rd_dias_config["sheets"]["included"],
                        "constant_fields": {"variant_category": "included"}}
        df = extract_tabular(rd_dias_cuh_workbook, "included", sheet_config, rd_dias_cuh_path)
        assert "variant_category" in df.columns
        assert (df["variant_category"] == "included").all()

    def test_multiple_constant_fields_all_added(self, rd_dias_cuh_workbook, rd_dias_cuh_path, rd_dias_config):
        """Multiple keys in constant_fields all appear as columns."""
        sheet_config = {**rd_dias_config["sheets"]["included"],
                        "constant_fields": {"variant_category": "included", "source_caller": "gatk"}}
        df = extract_tabular(rd_dias_cuh_workbook, "included", sheet_config, rd_dias_cuh_path)
        assert (df["variant_category"] == "included").all()
        assert (df["source_caller"] == "gatk").all()


class TestExtractTabularOptionalColumns:
    """Optional columns in tabular extraction."""

    def _make_workbook_and_path(self, tmp_path, columns: dict):
        """Write a minimal xlsx with given columns and return (workbook, path)."""
        import openpyxl
        df = pd.DataFrame(columns)
        path = tmp_path / "test.xlsx"
        df.to_excel(str(path), sheet_name="included", index=False)
        wb = openpyxl.load_workbook(str(path))
        return wb, path

    def test_optional_column_absent_is_skipped(self, tmp_path):
        """Optional column not in sheet does not raise; db_column absent from result."""
        wb, path = self._make_workbook_and_path(
            tmp_path, {"CHROM": ["1"], "HGVSc": ["c.100A>T"]}
        )
        sheet_config = {
            "extraction_type": "tabular",
            "columns": [
                {"source": "CHROM",  "db_column": "chromosome"},
                {"source": "HGVSc",  "db_column": "hgvsc"},
                {"source": "HGVSp",  "db_column": "hgvsp", "optional": True},
            ],
            "generated_columns": [],
        }
        df = extract_tabular(wb, "included", sheet_config, path)
        assert "chromosome" in df.columns
        assert "hgvsc" in df.columns
        assert "hgvsp" not in df.columns

    def test_optional_column_present_is_extracted(self, tmp_path):
        """Optional column present in sheet is extracted normally."""
        wb, path = self._make_workbook_and_path(
            tmp_path, {"CHROM": ["1"], "HGVSc": ["c.100A>T"], "HGVSp": ["p.Arg50Ter"]}
        )
        sheet_config = {
            "extraction_type": "tabular",
            "columns": [
                {"source": "CHROM",  "db_column": "chromosome"},
                {"source": "HGVSc",  "db_column": "hgvsc"},
                {"source": "HGVSp",  "db_column": "hgvsp", "optional": True},
            ],
            "generated_columns": [],
        }
        df = extract_tabular(wb, "included", sheet_config, path)
        assert "hgvsp" in df.columns
        assert df["hgvsp"].iloc[0] == "p.Arg50Ter"

    def test_required_column_absent_raises(self, tmp_path):
        """Required column (no optional flag) absent from sheet raises ValueError."""
        wb, path = self._make_workbook_and_path(
            tmp_path, {"CHROM": ["1"]}
        )
        sheet_config = {
            "extraction_type": "tabular",
            "columns": [
                {"source": "CHROM",  "db_column": "chromosome"},
                {"source": "HGVSc",  "db_column": "hgvsc"},  # required, absent
            ],
            "generated_columns": [],
        }
        with pytest.raises(ValueError):
            extract_tabular(wb, "included", sheet_config, path)


class TestParseWorkbookPindel:
    """parse_workbook pindel concatenation behaviour."""

    def _mock_wb(self, sheetnames):
        wb = MagicMock()
        wb.sheetnames = sheetnames
        return wb

    def _read_excel_side_effect(self, included_rows=3, pindel_rows=2):
        def _side_effect(file, sheet_name, usecols=None, nrows=None):
            n = included_rows if sheet_name == "included" else pindel_rows
            return pd.DataFrame({"CHROM": [str(i) for i in range(n)]})
        return _side_effect

    def test_pindel_rows_concatenated_with_included(self):
        wb = self._mock_wb(["included", "pindel"])
        with patch("augean.parser.pd.read_excel", side_effect=self._read_excel_side_effect(3, 2)):
            df = parse_workbook(wb, _MINIMAL_TWO_SHEET_CONFIG, Path("fake.xlsx"))
        assert len(df) == 5

    def test_variant_category_values_correct_per_source(self):
        wb = self._mock_wb(["included", "pindel"])
        with patch("augean.parser.pd.read_excel", side_effect=self._read_excel_side_effect(3, 2)):
            df = parse_workbook(wb, _MINIMAL_TWO_SHEET_CONFIG, Path("fake.xlsx"))
        assert set(df["variant_category"].unique()) == {"included", "pindel"}
        assert (df["variant_category"] == "included").sum() == 3
        assert (df["variant_category"] == "pindel").sum() == 2

    def test_pindel_absent_workbook_still_parses(self):
        """If pindel sheet is not in workbook.sheetnames, parsing continues without error."""
        wb = self._mock_wb(["included"])  # no pindel
        with patch("augean.parser.pd.read_excel", side_effect=self._read_excel_side_effect(3, 0)):
            df = parse_workbook(wb, _MINIMAL_TWO_SHEET_CONFIG, Path("fake.xlsx"))
        assert len(df) == 3
        assert (df["variant_category"] == "included").all()

    def test_top_level_constant_fields_present_on_all_rows(self):
        """Top-level constant_fields (e.g. allele_origin) broadcast to both included and pindel rows."""
        wb = self._mock_wb(["included", "pindel"])
        with patch("augean.parser.pd.read_excel", side_effect=self._read_excel_side_effect(3, 2)):
            df = parse_workbook(wb, _MINIMAL_TWO_SHEET_CONFIG, Path("fake.xlsx"))
        assert (df["allele_origin"] == "somatic").all()


# ---------------------------------------------------------------------------
# Smoke test: pindel sheet exists in test workbook
# ---------------------------------------------------------------------------

def test_haemonc_smoke_pindel_sheet_present():
    """Precondition: the canonical anonymised HaemOnc fixture has a pindel sheet."""
    import openpyxl
    fixture = WORKBOOKS_DIR / "haemonc" / "999999999-99999K9999-99NGSH999-9999-M-99999999.xlsx"
    if not fixture.exists():
        pytest.skip("Anonymised HaemOnc fixture not present")
    wb = openpyxl.load_workbook(fixture, read_only=True, data_only=True)
    try:
        assert "pindel" in wb.sheetnames, "Anonymised HaemOnc fixture must have a pindel sheet"
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Smoke tests: all HaemOnc workbooks in test_data parse without error
# ---------------------------------------------------------------------------

_HAEMONC_DIR = WORKBOOKS_DIR / "haemonc"
_HAEMONC_XLSX = sorted(_HAEMONC_DIR.glob("*.xlsx"))


def test_haemonc_smoke_inputs_present():
    assert _HAEMONC_XLSX, f"No .xlsx files found in {_HAEMONC_DIR}"


@pytest.mark.parametrize("xlsx_path", _HAEMONC_XLSX, ids=lambda p: p.name)
def test_haemonc_workbook_smoke(xlsx_path, haemonc_config):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    try:
        df = parse_workbook(wb, haemonc_config, xlsx_path)
    finally:
        wb.close()
    assert not df.empty
    assert "specimen_id" in df.columns
    assert "hgvsc" in df.columns
    assert "allele_origin" in df.columns
