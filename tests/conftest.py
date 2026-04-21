"""Shared test fixtures."""
import json
import re
from pathlib import Path
from unittest.mock import MagicMock

import openpyxl
import pytest

TEST_DATA_DIR = Path(__file__).parent / "test_data"
WORKBOOKS_DIR = TEST_DATA_DIR / "workbooks"
CONFIGS_DIR = Path(__file__).parent.parent / "configs"


# ---------------------------------------------------------------------------
# Acceptance test CLI options
# ---------------------------------------------------------------------------

def pytest_addoption(parser):
    parser.addoption(
        "--db_credentials",
        default=None,
        help="Path to db_credentials.json for acceptance tests",
    )
    parser.addoption(
        "--acceptance_schema",
        default="testdirectory",
        help="PostgreSQL schema to use for acceptance tests (default: testdirectory)",
    )


@pytest.fixture(scope="session")
def acceptance_db_credentials(request):
    """Path to DB credentials JSON for acceptance tests. Skips if not supplied."""
    creds_path = request.config.getoption("--db_credentials")
    if creds_path is None:
        pytest.skip("--db_credentials not provided; skipping acceptance test")
    return Path(creds_path)


@pytest.fixture(scope="session")
def acceptance_schema(request):
    schema = request.config.getoption("--acceptance_schema")
    if not re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", schema):
        pytest.fail(
            f"--acceptance_schema '{schema}' is not a valid PostgreSQL identifier. "
            "Use only letters, digits, and underscores, starting with a letter or underscore."
        )
    return schema


@pytest.fixture(scope="session")
def acceptance_engine(acceptance_db_credentials):
    """SQLAlchemy engine for acceptance tests."""
    from augean.db import create_engine as augean_create_engine
    with open(acceptance_db_credentials) as f:
        creds = json.load(f)
    return augean_create_engine(creds)


# ---------------------------------------------------------------------------
# Real workbook fixtures
# ---------------------------------------------------------------------------

def _first_xlsx(directory: Path):
    """Return the first .xlsx in directory, or None if absent."""
    files = sorted(directory.glob("*.xlsx"))
    return files[0] if files else None


_RD_DIAS_XLSX = _first_xlsx(WORKBOOKS_DIR / "rd_dias")
_HAEMONC_XLSX_PATH = _first_xlsx(WORKBOOKS_DIR / "haemonc")


@pytest.fixture(scope="session")
def rd_dias_cuh_path():
    if _RD_DIAS_XLSX is None:
        pytest.skip("No RD Dias test workbook available")
    return _RD_DIAS_XLSX


@pytest.fixture(scope="session")
def rd_dias_nuh_path():
    files = sorted((WORKBOOKS_DIR / "rd_dias").glob("*.xlsx"))
    if len(files) < 2:
        pytest.skip("No second RD Dias test workbook available")
    return files[1]


@pytest.fixture(scope="session")
def haemonc_path():
    if _HAEMONC_XLSX_PATH is None:
        pytest.skip("No HaemOnc test workbook available")
    return _HAEMONC_XLSX_PATH


@pytest.fixture(scope="session")
def rd_dias_cuh_workbook(rd_dias_cuh_path):
    return openpyxl.load_workbook(rd_dias_cuh_path, data_only=True)


@pytest.fixture(scope="session")
def rd_dias_nuh_workbook(rd_dias_nuh_path):
    return openpyxl.load_workbook(rd_dias_nuh_path, data_only=True)


@pytest.fixture(scope="session")
def haemonc_workbook(haemonc_path):
    return openpyxl.load_workbook(haemonc_path, data_only=True)


# ---------------------------------------------------------------------------
# Config fixtures
# ---------------------------------------------------------------------------

@pytest.fixture(scope="session")
def all_configs():
    from augean.config import load_configs
    return load_configs(CONFIGS_DIR)


@pytest.fixture(scope="session")
def rd_dias_config(all_configs):
    return next(c for c in all_configs if c["format_name"] == "rd_dias_v1")


@pytest.fixture(scope="session")
def haemonc_config(all_configs):
    return next(c for c in all_configs if c["format_name"] == "haemonc_uranus_v1")


# ---------------------------------------------------------------------------
# Mock workbook helper
# ---------------------------------------------------------------------------

def make_mock_workbook(sheetnames: list[str], cells: dict = None) -> MagicMock:
    """Create a minimal mock openpyxl Workbook.

    cells: dict of {sheet_name: {cell_address: value}}
    """
    wb = MagicMock()
    wb.sheetnames = sheetnames
    cells = cells or {}

    def get_sheet(name):
        sheet = MagicMock()
        sheet_cells = cells.get(name, {})

        def get_cell(addr):
            cell = MagicMock()
            cell.value = sheet_cells.get(addr)
            return cell

        # Simulate iter_rows for header detection
        def iter_rows(min_row=1, max_row=1, **kwargs):
            if min_row == 1 and max_row == 1:
                headers = sheet_cells.get("__headers__", [])
                mock_cells = []
                for h in headers:
                    mc = MagicMock()
                    mc.value = h
                    mock_cells.append(mc)
                yield mock_cells
        sheet.iter_rows = iter_rows

        # Column iteration (for sentinel_scan / label_scan)
        def col_iter(col):
            col_cells = sheet_cells.get(f"__col_{col}__", [])
            result = []
            for row_num, val in enumerate(col_cells, start=1):
                mc = MagicMock()
                mc.value = val
                mc.row = row_num
                result.append(mc)
            return result

        sheet.__getitem__ = lambda s, key: (
            col_iter(key) if (isinstance(key, str) and len(key) == 1 and key.isalpha() and not any(c.isdigit() for c in key))
            else get_cell(key)
        )
        return sheet

    wb.__getitem__ = lambda s, key: get_sheet(key)
    wb.__contains__ = lambda s, key: key in sheetnames
    return wb
