import logging
from pathlib import Path

import openpyxl

from hestia.config import get_config_for_workbook
from hestia.errors import AmbiguousWorkbookFormatError, WorkbookFormatUnknownError

log = logging.getLogger(__name__)


def load_workbook(path: Path) -> openpyxl.Workbook:
    """Open the xlsx file; raise descriptive error if unreadable."""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        log.debug("Opened workbook: %s (sheets: %s)", path.name, wb.sheetnames)
        return wb
    except Exception as exc:
        raise OSError(f"Cannot open workbook '{path}': {exc}") from exc


def detect_format(workbook, configs: list[dict]) -> dict:
    """Return the matched config for this workbook."""
    return get_config_for_workbook(workbook, configs)
